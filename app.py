"""
Founder Readiness Program — Assessment Tool
===========================================
Single-page questionnaire for founders (C1-C5).
Generates the standard .xlsx tracker without requiring a login.
"""

import streamlit as st
import io
import json
import os
from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
XLSX_TEMPLATE = "Tracker-Template.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Founder Readiness Assessment",
    page_icon="🚀",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# GLOBAL CSS
# ─────────────────────────────────────────────────────────────────────────────
def get_css(theme="Dark"):
    if theme == "Light":
        bg = "#f8fafc"
        sbg = "#ffffff"
        card = "#ffffff"
        text = "#000000"
        text_sub = "#475569"
        border = "rgba(15,23,42,0.1)"
        accent = "#f59e0b"  # Professional Gold/Amber
        hero = "linear-gradient(135deg, #f1f5f9 0%, #e2e8f0 100%)"
        shadow = "rgba(0,0,0,0.05)"
        btn_bg = "#ffffff"
        btn_text = "#000000"
    else:
        bg = "#070d19"
        sbg = "#0b1424"
        card = "#0f172a"
        text = "#f8fafc"
        text_sub = "#94a3b8"
        border = "rgba(255,255,255,0.08)"
        accent = "#e53935"  # Original Red
        hero = "linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%)"
        shadow = "rgba(0,0,0,0.5)"
        btn_bg = "#0f172a"
        btn_text = "#f8fafc"

    return f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');

    html, body, [class*="css"] {{
        font-family: 'Inter', sans-serif;
        color: {text};
        font-weight: 500;
    }}
    
    /* Ensure Streamlit icons don't break into text ligatures */
    span[class*="material"] {{
        font-family: 'Material Symbols Rounded', 'Material Icons' !important;
    }}

    .stApp {{ background: {bg}; }}
    .main .block-container {{ padding-top: 2rem; max-width: 1120px; }}

    section[data-testid="stSidebar"] {{
        background: {sbg};
        border-right: 1px solid {border};
    }}

    .hero-banner {{
        background: {hero};
        border-radius: 24px;
        padding: 2.5rem 3.5rem;
        margin-bottom: 2rem;
        border: 1px solid {border};
        box-shadow: 0 20px 50px {shadow};
        position: relative;
        overflow: hidden;
    }}
    .hero-banner h1 {{ color: {text}; font-size: 2.5rem; font-weight: 900; margin: 0; }}
    .hero-banner p {{ color: {text_sub}; font-size: 1.25rem; font-weight: 600; margin-top: 0.8rem; }}
    .hero-pill {{
        display: inline-block; background: {accent}; color: #fff;
        font-size: 0.85rem; font-weight: 800; padding: 5px 16px;
        border-radius: 20px; text-transform: uppercase; letter-spacing: 1px;
        margin-bottom: 1.2rem;
    }}

    .factor-card {{
        background: {card};
        border: 1px solid {border};
        border-radius: 16px;
        padding: 1.8rem;
        margin-bottom: 1.2rem;
        transition: transform 0.2s;
    }}
    .factor-card:hover {{ transform: translateY(-2px); border-color: {accent}; }}
    .factor-title {{ color: {text}; font-weight: 800; font-size: 1.25rem; margin-bottom: 0.6rem; }}
    .factor-question {{ color: {text_sub}; font-size: 1.05rem; line-height: 1.7; font-weight: 500; }}

    .prog-container {{
        background: rgba(255,255,255,0.03);
        border-radius: 12px;
        padding: 1.2rem;
        text-align: center;
        border: 1px solid {border};
    }}
    .prog-val {{ font-size: 2.6rem; font-weight: 900; color: {accent}; }}
    .prog-lbl {{ font-size: 0.8rem; color: {text_sub}; text-transform: uppercase; letter-spacing: 1px; font-weight: 700; }}

    /* Streamlit overrides */
    h1, h2, h3, h4, h5, h6, p, span, label, .stMarkdown {{
        color: {text} !important;
    }}
    
    div.stButton > button {{
        border-radius: 10px;
        font-weight: 700;
        background: {btn_bg};
        color: {btn_text};
        border: 1px solid {border};
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
    }}
    
    div.stButton > button:hover {{
        border-color: {accent};
        color: {accent};
        transform: translateY(-2px) scale(1.02);
        box-shadow: 0 10px 20px rgba(0,0,0,0.1);
    }}
    
    div.stButton > button:active {{
        transform: scale(0.98);
    }}

    /* Primary button override */
    div.stButton > button[kind="primary"] {{
        background: {accent} !important;
        color: #ffffff !important;
        border: none !important;
        box-shadow: 0 4px 15px {shadow};
    }}
    
    div.stButton > button[kind="primary"]:hover {{
        background: {accent} !important;
        opacity: 0.9;
        transform: translateY(-2px) scale(1.02);
        box-shadow: 0 8px 25px {shadow};
    }}

    /* Mobile Responsiveness */
    @media (max-width: 640px) {{
        .hero-banner {{
            padding: 1.5rem !important;
            border-radius: 16px !important;
        }}
        .hero-banner h1 {{
            font-size: 1.8rem !important;
        }}
        .main .block-container {{
            padding-left: 1rem !important;
            padding-right: 1rem !important;
        }}
        .factor-card {{
            padding: 1.2rem !important;
        }}
        .factor-title {{
            font-size: 1.1rem !important;
        }}
        /* Increase button height for easier tapping on mobile */
        div.stButton > button {{
            padding-top: 12px !important;
            padding-bottom: 12px !important;
            min-height: 50px !important;
        }}
    }}
    </style>
    """


# ─────────────────────────────────────────────────────────────────────────────
# DATA & HELPERS
# ─────────────────────────────────────────────────────────────────────────────

COHORTS = {
    "C1": {
        "name": "Validate & Launch",
        "anchor": "Product-Market Fit & Competitive Positioning",
        "domains": {
            "Product Maturity": {
                "weight": 0.35,
                "factors": [
                    ("Development stage",        "Which best describes your product today?\n1. Only an idea or concept — nothing built\n2. Prototype built but not tested outside my team\n3. MVP tested with at least 3 real users outside my team\n4. Product in real-world use with some paying users\n5. Product generating consistent revenue with multiple customers", 0.30, True),
                    ("External validation",       "How many people outside your team have used or tested your product?\n1. None\n2. 1–2 (family or friends only)\n3. 3–5 users from my target market\n4. 6–15 users from my target market\n5. More than 15 documented users from my target market", 0.25, True),
                    ("Paying users or pilots",    "Have any users paid for your product — even a small amount?\n1. No\n2. Not paid but we have a formal pilot agreement\n3. 1–2 paying users\n4. 3–10 paying users\n5. More than 10 paying users or a signed commercial contract", 0.20, False),
                    ("Documented feedback loops", "How do you capture and use customer feedback?\n1. I do not collect feedback formally\n2. I ask informally but do not record it\n3. I have basic records of what users said\n4. I document feedback and use it to make product decisions\n5. I have a regular feedback process that has driven at least 2 product changes", 0.15, False),
                    ("Technical feasibility",     "Are there unresolved technical risks that could stop your product?\n1. Yes — major risks unresolved\n2. Yes — some risks being worked on\n3. Minor risks only\n4. No significant risks — product works as intended\n5. No risks — product is proven and scalable", 0.05, False),
                    ("Differentiation",           "Can you state clearly why your product is better than existing alternatives?\n1. No — have not assessed competition\n2. I think it is better but cannot state why\n3. I have a reason but it is not backed by user evidence\n4. I have a clear differentiator confirmed by at least one user\n5. I have a documented advantage: IP, community access, process, or certified quality", 0.05, False),
                ],
            },
            "Compliance Readiness": {
                "weight": 0.20,
                "factors": [
                    ("Regulatory awareness",              "Do you know which regulatory approvals your product requires?\n1. No idea\n2. Aware approvals are needed but have not researched\n3. I know the specific approvals needed\n4. I know them and have started the process\n5. I have all required approvals OR my product does not require any", 0.20, False),
                    ("Industry compliance identified",    "Have you identified sector-specific compliance requirements (FSSAI, BIS, CDSCO, etc.)?\n1. No\n2. I think there are requirements but have not identified them\n3. I have identified the requirements\n4. I have identified them and have a plan to meet them\n5. All sector compliance requirements are met", 0.20, False),
                    ("Current compliance status",         "What is your current GST and ROC compliance status?\n1. Not registered / not compliant\n2. Registered but filing is irregular\n3. Registered and filing regularly\n4. Fully compliant with no pending obligations\n5. Fully compliant with a CA managing filings", 0.15, False),
                    ("Product documentation",             "Is your product documentation ready for regulatory inspection?\n1. No documentation exists\n2. Basic documentation started\n3. Documentation exists but incomplete\n4. Complete documentation ready for inspection\n5. Documentation reviewed and accepted by a regulatory body", 0.15, False),
                    ("Testing certifications",            "Have testing or certification processes been initiated?\n1. No\n2. Aware of requirements but not started\n3. Started the process\n4. Process underway — awaiting results\n5. Certifications obtained", 0.10, False),
                    ("Vendor and manufacturing compliance","Are your vendors aligned with the compliance standards your product requires?\n1. Not applicable or not considered\n2. Aware but have not verified\n3. Partially verified\n4. All main vendors verified\n5. Fully verified with documentation", 0.10, False),
                    ("Legal risk exposure",               "Are there any legal risks if your product were launched commercially today?\n1. Yes — significant legal risks\n2. Yes — some risks I am aware of\n3. Minor risks that can be managed\n4. No known legal risks\n5. No risks — legally verified", 0.05, False),
                    ("Revenue dependency on approvals",   "Can you generate revenue today without waiting for any regulatory approval?\n1. No — revenue fully blocked by approvals\n2. Partially — limited revenue possible\n3. Yes but constrained by current approval status\n4. Yes — revenue is not blocked\n5. Yes — fully cleared and scaling", 0.05, False),
                ],
            },
            "Market & Customer Clarity": {
                "weight": 0.35,
                "factors": [
                    ("Target user definition",         "How precisely can you describe your target customer?\n1. Broad category only\n2. Narrower category\n3. Specific profile (age, occupation, location, behaviour)\n4. Specific profile based on 5+ interviews\n5. Named individuals or organisations with intent to buy", 0.15, False),
                    ("Payer identification",            "Who pays for your product?\n1. Not clear — still working it out\n2. The user pays directly\n3. A third party pays — I know who\n4. Both user and third party — split clearly identified\n5. Payment model validated — I have received money from the actual payer", 0.10, False),
                    ("Influencer or decision-maker map","Have you identified the key decision-maker in the purchase of your product?\n1. No\n2. I think I know but have not verified\n3. I have identified them informally\n4. I have mapped decision-makers for my main customer segment\n5. I have named decision-makers with contact details and approach strategy", 0.10, False),
                    ("Problem validation",              "Have you validated that the problem your product solves is real and frequent?\n1. Not yet — assumed based on personal observation\n2. Discussed with 1–2 people informally\n3. Spoke with 3–5 potential customers outside my network\n4. Conducted structured interviews with 6+ potential customers\n5. Documented evidence from 10+ users including frequency and intensity data", 0.15, False),
                    ("Problem frequency and intensity", "Do you know how frequently the problem occurs and how painful it is?\n1. No — assumed\n2. Verbal impression only\n3. Basic understanding from a few conversations\n4. Quantified from structured interviews with 5+ users\n5. Rigorously documented with frequency data and intensity scores", 0.10, False),
                    ("Customer discovery volume",       "How many structured customer interviews have you completed?\n1. None\n2. 1–2\n3. 3–5\n4. 6–10\n5. More than 10", 0.15, False),
                    ("Insights documented",             "Are your customer discovery insights documented and informing decisions?\n1. No\n2. I have notes but have not reviewed them\n3. I review occasionally\n4. I refer to documented insights regularly\n5. Documented insights have driven at least 2 product or positioning decisions", 0.10, False),
                    ("Clear use case",                  "Is there one specific, demonstrable use case your product solves well?\n1. Not yet\n2. I have an idea but it is not specific\n3. One clear use case identified\n4. One use case demonstrated to a real user\n5. One use case proven with documented user evidence", 0.05, False),
                    ("Early traction or interest",      "Do you have documented signals of market demand?\n1. No\n2. Verbal interest only\n3. 1–2 written expressions of interest\n4. Signed LOI or pilot agreement\n5. Paid order or signed contract", 0.05, False),
                    ("Defined sales channel",           "Do you have a clear hypothesis for how your product reaches the customer?\n1. Not yet\n2. General idea (direct sales, online, etc.)\n3. Specific channel identified and tested once\n4. Channel tested with at least 3 successful transactions\n5. Channel is repeatable — more than 10 transactions through the same route", 0.05, False),
                ],
            },
            "Founder Readiness": {
                "weight": 0.10,
                "factors": [
                    ("Accurate self-assessment of stage","How well does your self-declared cohort (C1) match where your business actually is?\n1. I am probably in the wrong cohort — I have no product yet\n2. I may be in the wrong cohort — situation is unclear\n3. I am likely in the right cohort\n4. I am clearly in the right cohort and can explain why\n5. I am in the right cohort with evidence to prove it", 0.30, False),
                    ("Coachability",                    "When someone shows you evidence that contradicts an assumption, what do you do?\n1. I defend my assumption\n2. I listen but usually stick with my view\n3. I consider it but need time to change\n4. I update my view when the evidence is strong\n5. I actively seek out evidence that challenges my assumptions", 0.30, False),
                    ("Time available to execute",       "How much time can you dedicate to FRP activities between now and June 2026?\n1. Less than 5 hours per week\n2. 5–10 hours per week\n3. 10–20 hours per week\n4. More than 20 hours per week\n5. Full time — this is my only focus", 0.20, False),
                    ("Network for customer discovery",  "Do you have access to at least 10 potential customers you could speak to in the next 30 days?\n1. No — I do not have access to potential customers\n2. I can reach 1–3 people\n3. I can reach 4–9 people\n4. I can reach 10–20 people\n5. I can reach more than 20 potential customers within 30 days", 0.20, False),
                ],
            },
        },
    },
    "C2": {
        "name": "Grow & Scale D2C",
        "anchor": "Go-To-Market Execution",
        "domains": {
            "Founder & Business Profile": {
                "weight": 0.10,
                "factors": [
                    ("Founder background",           "How would you describe your background in sales, marketing, or D2C?\n1. No prior experience\n2. Some exposure but not formal\n3. I have run a D2C business or sales function before\n4. Significant experience — I led a team or function\n5. Deep expertise — D2C is my primary background", 0.30, False),
                    ("Reason for choosing the industry",  "Why did you choose this specific industry or product category?\n1. Circumstance — I fell into it\n2. Family or community tradition\n3. I saw a gap but have not validated it\n4. I validated a gap before starting\n5. Deep domain knowledge and validated gap — this is my advantage", 0.30, False),
                    ("USP clarity",                  "Can you state in one sentence why a consumer should choose your product?\n1. No — still working this out\n2. Yes but it is generic (better quality, good price)\n3. Yes — a specific functional reason\n4. Yes — specific reason confirmed by at least one customer\n5. Yes — specific, evidenced, already in use in marketing materials", 0.40, False),
                ],
            },
            "Revenue & Sales Performance": {
                "weight": 0.35,
                "factors": [
                    ("Current sales volume",     "What is your average monthly D2C revenue in the last 3 months (Rs)?\n1. Rs 0 — no sales yet\n2. Less than Rs 10,000\n3. Rs 10,000 to Rs 50,000\n4. Rs 50,000 to Rs 2,00,000\n5. More than Rs 2,00,000", 0.30, True),
                    ("Revenue consistency",      "How consistent is your D2C revenue month on month?\n1. No revenue\n2. Revenue occasionally — no pattern\n3. Some months have sales, some do not\n4. Revenue every month but varies significantly\n5. Revenue every month, growing or predictably stable", 0.20, False),
                    ("Product COGS and margin",  "Do you know your COGS per unit and your gross margin?\n1. No — not calculated\n2. Rough idea but not precise\n3. I know COGS but not the margin %\n4. I know both COGS and gross margin % for my main product\n5. I know COGS, margin, and break-even point for all key products", 0.20, True),
                    ("Average order value",      "What is the average value of a single customer order?\n1. Do not know\n2. Less than Rs 200\n3. Rs 200 to Rs 500\n4. Rs 500 to Rs 2,000\n5. More than Rs 2,000 or I actively manage AOV through bundles", 0.15, False),
                    ("Return rate",              "What is your approximate product return or refund rate?\n1. Do not track this\n2. More than 15%\n3. 10–15%\n4. Less than 10%\n5. Less than 5% with documented reasons for each return", 0.15, False),
                ],
            },
            "Marketing & Customer Acquisition": {
                "weight": 0.30,
                "factors": [
                    ("Active marketing channels",          "How many active marketing channels do you use consistently every week?\n1. None\n2. 1 channel\n3. 2 channels\n4. 3 channels\n5. 4+ channels with documented performance per channel", 0.20, False),
                    ("Customer acquisition cost",          "Do you know what it costs you to acquire one new customer?\n1. No\n2. Rough idea but no calculation\n3. Calculated CAC for one channel\n4. Know CAC for all active channels\n5. Know CAC, LTV, and the LTV:CAC ratio", 0.20, False),
                    ("Conversion rate",                    "Do you know what percentage of enquiries convert to paying customers?\n1. Do not track\n2. Less than 1%\n3. 1–3%\n4. 3–8%\n5. More than 8% with optimisation experiments underway", 0.15, False),
                    ("Repeat purchase rate",               "What percentage of your customers have bought from you more than once?\n1. Do not track\n2. Less than 10%\n3. 10–25%\n4. 25–50%\n5. More than 50% — I have a strong repeat customer base", 0.15, False),
                    ("Social media presence and strategy", "Do you post content on social media consistently for your business?\n1. No social media presence\n2. Occasional posts — no schedule\n3. At least 2–3 posts per week on one platform\n4. Consistent content on 2+ platforms with a theme\n5. Consistent strategy with engagement tracking and audience growth", 0.10, False),
                    ("Digital storefront readiness",       "How would you rate your online presence for converting visitors into buyers?\n1. No online presence\n2. Social media page only, no purchase mechanism\n3. WhatsApp Business or basic website — can receive orders\n4. Website or marketplace listing optimised for conversion\n5. Professional storefront with reviews, SEO, and conversion tracking", 0.10, False),
                    ("E-commerce platform presence",       "Is your product listed on any e-commerce platform?\n1. Not listed anywhere\n2. Considering but not started\n3. Listed but profile incomplete or inactive\n4. Active listing with at least 5 reviews\n5. Active listing with regular orders through the platform", 0.10, False),
                ],
            },
            "Operations & Fulfillment": {
                "weight": 0.15,
                "factors": [
                    ("Fulfillment process",          "How would you describe your order fulfillment process?\n1. No process — manage each order manually\n2. Basic process but not documented\n3. Documented process, can handle up to 20 orders/week\n4. Documented and tested, can scale to 100 orders/week\n5. Automated or semi-automated with quality checks", 0.30, False),
                    ("Delivery time and reliability","What is your average delivery time and how reliable is it?\n1. Do not track — variable\n2. More than 7 days, often missed\n3. 3–7 days, sometimes missed\n4. 2–4 days, rarely missed\n5. 1–2 days or same day, consistently met", 0.25, False),
                    ("Inventory management",         "How do you manage your inventory?\n1. No system — ad hoc\n2. Basic manual tracking\n3. Spreadsheet-based with regular review\n4. Systematic tracking with reorder triggers\n5. Automated or semi-automated inventory management", 0.25, False),
                    ("Capacity to scale",            "If your orders tripled next month, could you fulfill them?\n1. No — I would be completely overwhelmed\n2. Unlikely — I have no spare capacity\n3. Possibly — with significant strain\n4. Yes — with some planning\n5. Yes — I have headroom and a scale plan", 0.20, False),
                ],
            },
            "Pricing & Competitive Position": {
                "weight": 0.10,
                "factors": [
                    ("Price competitiveness", "How does your price compare to the nearest competitor?\n1. Do not know who my competitors are\n2. My price is higher with no clear reason\n3. My price is similar — competing on price\n4. My price is positioned for a specific reason and I can explain it\n5. Pricing strategy is deliberate, documented, and tested with customers", 0.40, False),
                    ("4P attributes",         "Have you thought through Product, Price, Place (channel), and Promotion as a coherent strategy?\n1. No — have not considered this\n2. Aware of the 4Ps but have not applied them\n3. Applied the framework informally\n4. Documented 4P strategy for my main product\n5. Documented, tested, and actively optimised 4P strategy", 0.35, False),
                    ("Number of market players","How well do you know the competitive landscape?\n1. I do not know who my competitors are\n2. I know a few names but no detail\n3. I have identified 3–5 main competitors and know their positioning\n4. I have a detailed competitive map with pricing, channels, and USP\n5. I track competitor movements regularly and adjust my strategy", 0.25, False),
                ],
            },
        },
    },
    "C3": {
        "name": "Acquire & Expand B2B",
        "anchor": "B2B Sales, Government Procurement & GeM Portal",
        "domains": {
            "B2G Sub-Track": {
                "weight": 0.10,
                "factors": [
                    ("GeM portal registration status",          "What is your GeM portal status?\n1. Not registered\n2. Registration started but not complete\n3. Registered but profile not fully set up\n4. Registered with at least 1 product listed\n5. Active seller — have received or bid on at least 1 GeM order", 0.25, False),
                    ("Tender eligibility awareness",             "Do you know which government tender categories you are eligible for?\n1. No\n2. Aware tenders exist but have not checked\n3. Identified 1–2 relevant categories\n4. Identified categories and have Udyam/NSIC registration\n5. Have applied to at least 1 tender", 0.25, False),
                    ("Documentation readiness",                  "Do you have the standard tender documents ready?\n1. Most documents missing\n2. About half ready\n3. Most ready, 1–2 missing\n4. All ready but not organised\n5. All ready and organised in a physical or digital data room", 0.20, False),
                    ("Past performance record",                  "Do you have any documented record of prior government supply?\n1. None\n2. Informal supply with no documentation\n3. 1 documented government transaction\n4. 2–3 documented government transactions\n5. 4+ with completion certificates or delivery records", 0.20, False),
                    ("SC/ST procurement preference awareness",   "Are you aware of SC/ST procurement preferences under NSIC, GeM, and PPP 2012?\n1. Not aware\n2. Heard about it but do not know details\n3. Know the policy but have not used it\n4. Know the policy and have used it in at least one bid\n5. Actively use SC/ST preferences in all eligible bids", 0.10, False),
                ],
            },
            "Market & Sales Readiness": {
                "weight": 0.35,
                "factors": [
                    ("POC stage",                 "Have you completed a proof of concept with a real B2B or government client?\n1. No B2B engagement yet\n2. Informal discussion only\n3. Completed an unpaid pilot or trial\n4. Completed a paid pilot with documented outcomes\n5. Have at least 1 paying B2B client on a formal contract", 0.20, True),
                    ("Product or service status", "Is your product ready to deliver at the scale of a typical B2B contract?\n1. Still in development\n2. Works but cannot handle a large order\n3. Can handle small contracts (up to Rs 5L)\n4. Can handle medium contracts (up to Rs 25L)\n5. Proven at contract scale", 0.10, False),
                    ("Piloting completion",        "Have you completed formal pilot engagements?\n1. No pilots done\n2. 1 informal trial with no documentation\n3. 1 formal pilot with some documentation\n4. 2+ pilots with documented outcomes\n5. 3+ pilots with at least 1 converted to a paid engagement", 0.15, False),
                    ("Geographical reach",         "How many districts or states does your business currently serve?\n1. Only within my own district\n2. 2–3 districts\n3. Across Tamil Nadu\n4. Multiple states\n5. National or international reach", 0.05, False),
                    ("Ideal customer profile",     "How precisely have you defined your target B2B buyer?\n1. No specific target\n2. Broad category\n3. Specific industry and company size\n4. Specific industry, size, and decision-maker role\n5. Named target accounts with contact details and engagement history", 0.10, False),
                    ("Marketing team",             "Do you have dedicated sales or marketing capacity?\n1. None — I handle everything\n2. I do sales alongside operations\n3. Part-time sales or marketing person\n4. Full-time sales or marketing person\n5. Dedicated sales team with pipeline management", 0.05, False),
                    ("Advertising budget and mix", "Do you have a dedicated marketing or advertising budget?\n1. No budget\n2. Ad hoc spend only\n3. Fixed monthly budget but no channel strategy\n4. Fixed budget allocated across channels\n5. Optimised budget with performance tracking per channel", 0.05, False),
                    ("Social media strategy",      "Do you have an active LinkedIn or B2B social media presence?\n1. No presence\n2. Personal profile only, no business content\n3. Business page exists but rarely updated\n4. Regular content targeting B2B decision-makers\n5. Active strategy with documented engagement and lead generation", 0.05, False),
                    ("Trial marketing and pipeline","Do you have an active B2B sales pipeline with named prospects?\n1. No prospects identified\n2. 1–2 informal conversations\n3. 3–5 named prospects in active discussion\n4. 5–10 prospects with documented pipeline value\n5. More than 10 qualified prospects with documented follow-up schedule", 0.05, False),
                    ("Price competitiveness",       "How does your pricing compare to B2B alternatives?\n1. Do not know\n2. Significantly more expensive\n3. Comparable — no price advantage\n4. Competitive on price\n5. Best value proposition — documented from buyer conversations", 0.05, False),
                    ("Number of market players",    "How well do you know your B2B competitive landscape?\n1. No knowledge of competitors\n2. I know a few names\n3. I have identified 3–5 main competitors\n4. I have a detailed competitive map\n5. I track competitor activity and adjust my B2B positioning accordingly", 0.05, False),
                    ("USP for B2B buyers",          "Can you state the specific business value your product creates for a B2B buyer?\n1. No — still working this out\n2. Generic statement\n3. Specific functional benefit\n4. Specific benefit confirmed by at least one B2B buyer\n5. Specific, quantified benefit confirmed by 3+ B2B buyers", 0.05, False),
                    ("4P attributes",               "Have you thought through your B2B positioning across Product, Price, Channel, and Promotion?\n1. Not considered this\n2. Aware of 4Ps but not applied\n3. Applied informally\n4. Documented 4P B2B positioning\n5. Tested and optimised B2B 4P positioning", 0.05, False),
                ],
            },
            "Financial Health": {
                "weight": 0.30,
                "factors": [
                    ("Turnover",                         "What is your average monthly revenue over the last 12 months (Rs)?\n1. Less than Rs 50,000\n2. Rs 50,000 to Rs 2,00,000\n3. Rs 2,00,000 to Rs 10,00,000\n4. Rs 10,00,000 to Rs 50,00,000\n5. More than Rs 50,00,000", 0.25, True),
                    ("Profit and loss position",         "Is your business currently profitable?\n1. Operating at a significant loss\n2. Operating at a small loss\n3. Breaking even\n4. Profitable — positive net margin\n5. Profitable with EBITDA margin above 15%", 0.20, False),
                    ("Initial funding and capital structure","How has your business been primarily funded?\n1. No formal funding\n2. Personal savings only\n3. Family and friends funding\n4. Government grant or scheme\n5. External investment (angel, debt, equity)", 0.10, False),
                    ("Expenditure trend",                "How is your cost base trending relative to your revenue?\n1. Costs are rising much faster than revenue\n2. Costs are rising faster than revenue\n3. Costs and revenue are growing at similar rates\n4. Revenue is growing faster than costs\n5. Revenue is growing significantly faster — improving margins", 0.15, False),
                    ("Cost per employee",                "Do you know your revenue per employee and cost per employee?\n1. No — have not calculated\n2. Rough estimate only\n3. Calculated but not benchmarked\n4. Calculated and benchmarked\n5. Calculated, benchmarked, and used to make hiring decisions", 0.10, False),
                    ("Promoter equity and compensation",  "Is your promoter equity clearly documented and your own compensation structured?\n1. No formal documentation\n2. Informal arrangement only\n3. Documented but not formalised\n4. Documented and formalised by CA or lawyer\n5. Documented, formalised, and transparent for investor due diligence", 0.05, False),
                    ("Return on investment",             "Do you track the return on your major expenditure items?\n1. No — spend without measuring outcome\n2. Track revenue but not ROI per item\n3. Track ROI informally for major items\n4. Track ROI for all major expenditure categories\n5. Track ROI with documented evidence and use it to reallocate spend", 0.10, False),
                    ("Authorised and paid-up capital",   "Do you know your company's authorised and paid-up capital?\n1. Not aware of these terms\n2. Aware but do not know the numbers\n3. Know the numbers but they are outdated\n4. Know the current numbers\n5. Know the numbers, current, and understand headroom for new investment", 0.05, False),
                ],
            },
            "Operational & Manufacturing Capability": {
                "weight": 0.25,
                "factors": [
                    ("Development stage",                       "Is your product ready to deliver at the scale of a typical government or B2B contract?\n1. Product still in development\n2. Product works but cannot handle a large order\n3. Can handle small contracts (up to Rs 5L)\n4. Can handle medium contracts (up to Rs 25L)\n5. Proven at scale — have delivered at that level", 0.25, True),
                    ("Factory or facility availability",        "Do you have your own production or service delivery facility?\n1. No facility — production is ad hoc\n2. Using shared or informal space\n3. Rented or owned but under-equipped\n4. Own or rented facility adequate for current scale\n5. Own facility with capacity to scale and documented quality processes", 0.20, False),
                    ("Know-how and innovation",                 "Does your business have proprietary know-how that competitors cannot easily replicate?\n1. No proprietary advantage\n2. Some differentiated knowledge but not formalised\n3. Documented process advantage\n4. Proprietary process or technology with some IP protection\n5. Patented or trade-secret protected advantage", 0.15, False),
                    ("Resources - machines, materials, people","Are your resources adequate to fulfil contract volumes you are targeting?\n1. Significant gaps in machines, materials, or people\n2. Major gaps in 1–2 areas\n3. Minor gaps — can be filled quickly\n4. Resources adequate for near-term contracts\n5. Resources optimised with clear scale-up plan", 0.15, False),
                    ("Outsourcing dependencies",                "Are you critically dependent on any single outsourced partner?\n1. Yes — entire production is outsourced with no backup\n2. Yes — critical dependency with no verified backup\n3. Some dependency but alternatives identified\n4. Low dependency — multiple supplier options\n5. No critical outsourcing dependencies", 0.10, False),
                    ("Advantage of location",                   "Does your business location give you a competitive advantage?\n1. Location is a disadvantage\n2. No advantage or disadvantage\n3. Some logistical benefit\n4. Clear location advantage for reaching buyers or clusters\n5. Location advantage is a documented part of the competitive positioning", 0.10, False),
                    ("Tax holidays, subsidies, IT adaptability","Is your business leveraging available government incentives, subsidies, or tax benefits?\n1. Not aware of what is available\n2. Aware but not yet applied\n3. Applied for some benefits\n4. Actively using at least 2 government benefits\n5. Full utilisation of available benefits with IT and compliance systems in place", 0.05, False),
                ],
            },
        },
    },
    "C4": {
        "name": "Fundraise & Accelerate",
        "anchor": "Investment Readiness Boot Camp & Financial Statements Decoded",
        "domains": {
            "Financial Documentation": {
                "weight": 0.40,
                "factors": [
                    ("3 years audited financial statements", "Do you have CA-signed audited financial statements for the last 3 years?\n1. No audits done\n2. 1 year audited\n3. 2 years audited\n4. 3 years audited but not organised\n5. 3 years audited, CA-signed, in a data room ready to share", 0.25, True),
                    ("Current year provisional statements",  "Do you have current year provisional or unaudited financial statements?\n1. No\n2. I have basic records but not formal statements\n3. Provisional P&L prepared\n4. Provisional P&L and balance sheet prepared\n5. Full provisional financials prepared by CA", 0.15, False),
                    ("1 year bank statements",               "Do you have 12 consecutive months of bank statements available?\n1. No\n2. Some months but not all 12\n3. 12 months but from multiple accounts not consolidated\n4. 12 months from primary account\n5. 12 months from all accounts, reconciled to P&L", 0.20, False),
                    ("Cap table",                            "Do you have a current cap table showing all shareholders and ownership percentages?\n1. No — have not thought about this\n2. I know roughly who owns what but it is not documented\n3. Basic cap table exists but not verified\n4. Cap table exists, CA-verified, includes convertible instruments\n5. Current, CA-verified cap table in data room with vesting schedules if applicable", 0.20, True),
                    ("Debentures and convertible instruments","Do you have any existing debentures, convertible notes, or SAFE instruments?\n1. None and not aware of what these are\n2. None but I understand what they are\n3. None — clean cap table\n4. Some instruments but documentation is informal\n5. All instruments formally documented and listed in data room", 0.10, False),
                    ("2 years financial projections",        "Do you have a financial model with 2-year projections?\n1. No projections\n2. Revenue projections only — no cost model\n3. Revenue and cost projections — no cash flow\n4. Full P&L projections with stated assumptions\n5. Full P&L and cash flow projections with scenario analysis and milestone linkage", 0.10, False),
                ],
            },
            "Traction & Business Fundamentals": {
                "weight": 0.25,
                "factors": [
                    ("Recurring revenue",           "Do you have recurring revenue (MRR or ARR)?\n1. No revenue\n2. Occasional sales — no recurring element\n3. Some recurring customers but not formalised\n4. Formalised recurring revenue: subscriptions, retainer, or repeat orders on contract\n5. Documented MRR/ARR that has grown for at least 3 consecutive months", 0.30, True),
                    ("Unit economics",              "Do you know your CAC, LTV, gross margin, and burn rate?\n1. No — not calculated\n2. I know 1 of these 4 numbers\n3. I know 2–3 of these numbers\n4. I know all 4 for my primary product or segment\n5. I know all 4 and can show the trend over the last 6 months", 0.25, False),
                    ("Customer concentration risk", "What percentage of your revenue comes from your top 1–2 customers?\n1. More than 80%\n2. 60–80%\n3. 40–60%\n4. 20–40%\n5. Less than 20% — well diversified", 0.15, False),
                    ("Competitive moat",            "What is your primary competitive moat?\n1. No clear moat\n2. Some differentiation but easily replicated\n3. One defensible advantage (community trust, location, IP, process)\n4. Two documented advantages\n5. Multiple reinforcing moats with evidence that they compound over time", 0.20, False),
                    ("Team completeness",           "Does your team have the capability to deploy investment capital and execute the scale plan?\n1. Team is founder only with no plan to hire\n2. Founder plus 1–2 people but critical gaps exist\n3. Most capabilities covered — 1–2 key hires needed\n4. Team is largely complete — key hires identified and planned\n5. Full team in place with all capabilities required for the next growth phase", 0.10, False),
                ],
            },
            "Investment Readiness": {
                "weight": 0.25,
                "factors": [
                    ("Valuation basis",                  "Do you have a defensible basis for your target valuation?\n1. No — have not thought about valuation\n2. A number in mind but no methodology\n3. Used a comparable company approach\n4. Used revenue multiple or DCF with stated assumptions\n5. Valuation memo prepared with 2+ methodologies", 0.20, False),
                    ("Use of funds clarity",             "Can you break down exactly how you will use the investment you are seeking?\n1. No — general idea only\n2. I know the broad categories\n3. Breakdown by category but no milestone linkage\n4. Breakdown by category linked to specific milestones\n5. Detailed breakdown, milestone-linked, with timeline and accountability", 0.20, False),
                    ("Understanding of dilution",        "Do you understand what dilution means and how it affects your ownership?\n1. Not sure what dilution means\n2. I know the concept but not how to calculate it\n3. I can calculate basic dilution\n4. I understand dilution across rounds including option pool expansion\n5. I can model dilution across multiple rounds and negotiate accordingly", 0.15, False),
                    ("SAFE and convertible note literacy","Do you understand how SAFE notes and convertible instruments work?\n1. Not sure what a SAFE note is\n2. I have heard of them but do not understand the mechanics\n3. I understand the basics\n4. I can explain how a SAFE converts and what the cap and discount mean\n5. I have structured or negotiated SAFE or convertible instruments before", 0.15, False),
                    ("Pitch deck quality",               "How complete and investor-ready is your current pitch deck?\n1. No deck\n2. Basic slides only — not structured as an investor deck\n3. Structured deck but missing 3+ key sections\n4. Complete deck with all sections but not yet investor-polished\n5. Deck reviewed by at least 1 investor or advisor with feedback incorporated", 0.15, False),
                    ("Data room completeness",           "Do you have an organised investor data room?\n1. No data room\n2. A folder with some documents but not organised\n3. Most documents present but disorganised\n4. Complete data room with all 8 standard documents\n5. Complete, organised data room shared with at least 1 investor for review", 0.15, False),
                ],
            },
            "Investor Engagement Readiness": {
                "weight": 0.10,
                "factors": [
                    ("Investor pipeline",                 "Do you have a list of target investors with warm or cold contact status?\n1. No investor list\n2. A few names but no contact strategy\n3. A list of 5–10 investors with categorisation\n4. Active list of 10+ investors with contact history\n5. Pipeline of 15+ investors with documented outreach, response status, and next steps", 0.30, False),
                    ("Introduction readiness",            "If a warm introduction to an investor happened tomorrow, are you ready?\n1. No — I would not know what to say\n2. I have a rough pitch but no specific ask\n3. I have a clear ask but no one-liner or memo\n4. I have a clear ask, a one-liner, and a context memo ready\n5. I have been introduced to at least 1 investor and know exactly what the next steps are", 0.25, False),
                    ("Due diligence preparedness",        "If an investor requested full due diligence tomorrow, could you respond within 48 hours?\n1. No — major documents are missing\n2. Some documents ready but gaps are significant\n3. Most documents ready — 2–3 items missing\n4. All documents ready with minor gaps\n5. Fully prepared — data room is complete and has been reviewed", 0.25, False),
                    ("Post-funding governance awareness", "Do you understand the governance obligations that come with accepting external investment?\n1. Not aware of post-investment obligations\n2. Aware they exist but do not know the details\n3. Basic understanding of board rights and reporting\n4. Can explain key investor rights and reporting requirements\n5. Have reviewed or negotiated a term sheet and understand all governance provisions", 0.20, False),
                ],
            },
        },
    },
    "C5": {
        "name": "Debt & Non-Dilutive",
        "anchor": "Investment Readiness Boot Camp & Financial Statements Decoded",
        "domains": {
            "Creditworthiness & Loan Readiness": {
                "weight": 0.35,
                "factors": [
                    ("Credit score and history",    "What is your current CIBIL or Experian credit score?\n1. I have not checked\n2. Below 600\n3. 600–699\n4. 700–749\n5. 750 or above with no defaults in the last 2 years", 0.25, True),
                    ("Consistent bank inflows",     "How regular are the cash inflows into your business bank account?\n1. No regular inflows\n2. Inflows are irregular — some months very low\n3. Some regularity but significant variation\n4. Regular inflows every month for the last 6+ months\n5. Consistent, growing inflows every month for 12+ months with no months below Rs 50,000", 0.30, True),
                    ("Existing debt obligations",   "Do you have any existing loans or EMI obligations?\n1. No loans of any kind\n2. 1 personal loan or consumer loan\n3. 1–2 business loans currently active\n4. Multiple loans but all current (no defaults)\n5. Multiple loans, all current, with documented repayment history", 0.15, False),
                    ("Collateral availability",     "Do you have any assets that could serve as collateral?\n1. No assets of any kind\n2. Personal property but cannot pledge it\n3. Own property or vehicle that could be pledged\n4. Business machinery or equipment that could be pledged\n5. Combination of assets sufficient to collateralise a loan of Rs 25L+", 0.15, False),
                    ("CGTMSE eligibility",          "Are you MSME-registered and aware of CGTMSE collateral-free loan coverage?\n1. Not MSME registered\n2. MSME registered but not aware of CGTMSE\n3. MSME registered, aware of CGTMSE but not applied\n4. Applied for CGTMSE-covered loan but not yet approved\n5. Have an active CGTMSE-covered loan or have successfully used the scheme", 0.10, False),
                    ("Repayment plan realism",      "Have you calculated whether your projected cash flows can service a loan EMI?\n1. No — not calculated\n2. Rough estimate only\n3. Calculated for one loan scenario\n4. Calculated for 2–3 scenarios\n5. Calculated with sensitivity analysis — DSCR above 1.25 for the target loan amount", 0.05, False),
                ],
            },
            "Business Fundamentals for Debt": {
                "weight": 0.15,
                "factors": [
                    ("Revenue consistency",              "Has your business generated revenue in every month for the last 6 months?\n1. No revenue in the last 6 months\n2. Revenue in some months only — no pattern\n3. Revenue most months but with 1–2 zero months\n4. Revenue every month for 6+ months\n5. Revenue every month for 12+ months, growing or stable", 0.35, True),
                    ("Business age and track record",    "How long has your business been operating with formal registration and records?\n1. Less than 6 months\n2. 6–12 months\n3. 1–2 years\n4. 2–3 years\n5. More than 3 years with unbroken records", 0.25, False),
                    ("Purpose of loan",                  "What will you primarily use the loan for?\n1. To cover ongoing operating losses\n2. To pay off personal debts\n3. For a mix of personal and business needs\n4. For working capital or equipment — specific business purpose\n5. For a specific, documented growth investment with projected ROI", 0.25, False),
                    ("Promoter personal financial health","How would you describe your personal financial health?\n1. Significant personal debts or defaults\n2. Some personal financial challenges\n3. Personal finances are neutral\n4. Personal finances are clean and organised\n5. Personal finances are clean, well-documented, and separated from business finances", 0.15, False),
                ],
            },
            "Non-Dilutive Scheme Awareness": {
                "weight": 0.15,
                "factors": [
                    ("Scheme eligibility mapping",           "Have you identified which government schemes you are eligible for?\n1. No\n2. Aware schemes exist but have not checked eligibility\n3. Checked eligibility for 1–2 schemes\n4. Identified all relevant schemes and eligibility criteria\n5. Have active applications in at least 1 scheme", 0.30, False),
                    ("Application documentation readiness",  "Do you have all documents needed for a loan or scheme application?\n1. Most documents missing\n2. About half ready\n3. Most ready, 1–2 missing\n4. All ready but not in one place\n5. All ready, organised, copied, and in a physical or digital file", 0.30, False),
                    ("Grant and subsidy awareness",          "Are you aware of grants or subsidies available to SC/ST entrepreneurs that do not require repayment?\n1. No awareness\n2. Heard about them generally\n3. Know 1–2 specific grants or subsidies with eligibility criteria\n4. Know 3+ grants or subsidies with eligibility criteria\n5. Applied for at least 1 grant or subsidy", 0.20, False),
                    ("Prior scheme utilisation",             "Have you already used any government scheme or benefited from any subsidy?\n1. No — never used any scheme\n2. Applied but not received\n3. Received a minor benefit (registration fee waiver, etc.)\n4. Used 1 scheme with documented outcome\n5. Used 2+ schemes with documented outcomes and compliance", 0.20, False),
                ],
            },
            "Financial Documentation": {
                "weight": 0.35,
                "factors": [
                    ("3 years audited financial statements",     "Do you have CA-signed audited financial statements for the last 3 years?\n1. No audits done\n2. 1 year audited\n3. 2 years audited\n4. 3 years audited but not organised\n5. 3 years audited, CA-signed, in a data room ready to share", 0.25, False),
                    ("Current year provisional statements",      "Do you have current year provisional or unaudited financial statements?\n1. No\n2. Basic records but not formal\n3. Provisional P&L prepared\n4. Provisional P&L and balance sheet prepared\n5. Full provisional financials prepared by CA", 0.15, False),
                    ("1 year bank statements",                   "Do you have 12 consecutive months of bank statements available?\n1. No\n2. Some months but not all 12\n3. 12 months but from multiple accounts not consolidated\n4. 12 months from primary account\n5. 12 months from all accounts, reconciled to P&L", 0.25, False),
                    ("Cap table",                                "Do you have a current cap table?\n1. No — have not thought about this\n2. I know roughly who owns what but it is not documented\n3. Basic cap table exists but not verified\n4. Cap table exists, CA-verified\n5. Current, CA-verified cap table in data room", 0.15, False),
                    ("Debentures and existing debt instruments",  "Do you have any existing debentures, convertible notes, or SAFE instruments?\n1. None and not aware of what these are\n2. None but I understand what they are\n3. None — clean cap table\n4. Some instruments but documentation is informal\n5. All instruments formally documented", 0.10, False),
                    ("2 years financial projections",            "Do you have a financial model with 2-year projections?\n1. No projections\n2. Revenue projections only\n3. Revenue and cost projections — no cash flow\n4. Full P&L projections with stated assumptions\n5. Full P&L and cash flow projections with scenario analysis", 0.10, False),
                ],
            },
        },
    },
}

SCORE_LABELS = {1: "Not Present", 2: "Aware", 3: "In Progress", 4: "Established", 5: "Strong"}

def init_session():
    if "scores" not in st.session_state: st.session_state.scores = {}
    if "remarks" not in st.session_state: st.session_state.remarks = {}
    if "active_cohort" not in st.session_state: st.session_state.active_cohort = "C1"
    if "theme" not in st.session_state: st.session_state.theme = "Dark"
    if "info" not in st.session_state: st.session_state.info = {
        "startup": "", "founder": "", "week": "Week 1"
    }

def build_xlsx(info, scores, remarks):
    with open(XLSX_TEMPLATE, "rb") as f:
        buf = io.BytesIO(f.read())
    wb = load_workbook(buf)
    
    # Update Startup_Info
    if "Startup_Info" in wb.sheetnames:
        si = wb["Startup_Info"]
        si["B3"] = info["startup"]
        # Find Founder Name label
        for row in si.iter_rows(min_row=1, max_row=10, max_col=2):
            if str(row[0].value).strip() == "Founder Name":
                row[1].value = info["founder"]

    # Update Weekly_Input
    if "Weekly_Input" in wb.sheetnames:
        wi = wb["Weekly_Input"]
        # Assuming Week 1 is Column H (8th column)
        for r_idx, row in enumerate(wi.iter_rows(min_row=6), start=6):
            c_val = row[0].value
            d_val = row[1].value
            f_val = row[3].value
            if not c_val: continue
            key = f"{c_val}|{d_val}|{f_val}"
            if key in scores:
                wi.cell(row=r_idx, column=8).value = scores[key]
                rk_key = key + "_remark"
                if remarks.get(rk_key):
                    wi.cell(row=r_idx, column=11).value = remarks[rk_key]
                    
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────────────────────────────────────

def main():
    init_session()
    st.markdown(get_css(st.session_state.theme), unsafe_allow_html=True)
    
    # ── Sidebar ──
    with st.sidebar:
        st.markdown("<h1 style='text-align:center;'>🚀</h1>", unsafe_allow_html=True)
        st.markdown("<h2 style='text-align:center; color:#e53935; margin-top:0;'>FRP Tracker</h2>", unsafe_allow_html=True)
        
        # Theme Toggle
        theme = st.radio("🎨 App Theme", ["Dark", "Light"], index=0 if st.session_state.theme == "Dark" else 1, horizontal=True)
        if theme != st.session_state.theme:
            st.session_state.theme = theme
            st.rerun()
        
        st.divider()
        
        st.markdown("### 🏢 Startup Details")
        st.session_state.info["startup"] = st.text_input("Startup Name", value=st.session_state.info["startup"])
        st.session_state.info["founder"] = st.text_input("Founder Name", value=st.session_state.info["founder"])
        st.session_state.info["week"]    = st.selectbox("Current Week", ["Week 1", "Week 2", "Week 3", "Week 4"], index=0)
        
        st.divider()
        if st.button("🗑 Reset All Data", use_container_width=True):
            st.session_state.scores = {}
            st.session_state.remarks = {}
            st.rerun()
            
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.info("Fill all sections to generate your professional readiness report.")

    # ── Hero ──
    st.markdown(f"""
    <div class="hero-banner">
        <div class="hero-pill">Bootcamp 2026 · Assessment</div>
        <h1>Founder Readiness Program</h1>
        <p>A comprehensive assessment for TN SC/ST Startup Fund candidates.</p>
    </div>
    """, unsafe_allow_html=True)

    # ── Cohort Selector ──
    st.markdown("### 🎯 Select Assessment Phase")
    st.markdown("<p style='margin-top:-15px; margin-bottom:15px; opacity:0.8; font-size:0.95rem;'>Click any phase below to switch; <b>all data is saved automatically with no loss.</b></p>", unsafe_allow_html=True)
    c_cols = st.columns(len(COHORTS))
    for i, (c_id, c_info) in enumerate(COHORTS.items()):
        # Progress calculation
        c_keys = [f"{c_id}|{d}|{f[0]}" for d, dv in c_info["domains"].items() for f in dv["factors"]]
        c_filled = sum(1 for k in c_keys if k in st.session_state.scores)
        c_total = len(c_keys)
        c_pct = int(c_filled / c_total * 100) if c_total else 0
        
        is_active = (st.session_state.active_cohort == c_id)
        if c_cols[i].button(f"{c_id}\n{c_pct}%", key=f"btn_c_{c_id}", use_container_width=True, type="primary" if is_active else "secondary"):
            st.session_state.active_cohort = c_id
            st.rerun()

    active_c = st.session_state.active_cohort
    c_data = COHORTS[active_c]
    
    st.markdown(f"""
    <div style='background:rgba(229,57,53,0.1); border-left:4px solid #e53935; padding:1.2rem; border-radius:8px; margin:1.5rem 0;'>
        <h3 style='margin:0; color:#fff;'>{active_c}: {c_data['name']}</h3>
        <p style='margin:0.3rem 0 0; color:rgba(255,255,255,0.6);'>{c_data['anchor']}</p>
    </div>
    """, unsafe_allow_html=True)

    # ── Factor Form ──
    q_idx = 1
    for d_name, d_info in c_data["domains"].items():
        st.markdown(f"#### 📁 {d_name}")
        for f in d_info["factors"]:
            f_name, q_text, f_w, is_hd = f
            key = f"{active_c}|{d_name}|{f_name}"
            
            st.markdown(f"""
            <div class="factor-card">
                <div class="factor-title">Qno.{q_idx}: {f_name} {'<span style="background:#b71c1c;color:#fff;font-size:0.68rem;font-weight:700;padding:1px 6px;border-radius:4px;">HD</span>' if is_hd else ''}</div>
                <div class="factor-question">{q_text.replace('\n','<br>')}</div>
            </div>
            """, unsafe_allow_html=True)
            q_idx += 1
            
            cur_s = st.session_state.scores.get(key)
            sc_cols = st.columns(5)
            for sc_idx, (sv, sl) in enumerate(SCORE_LABELS.items()):
                with sc_cols[sc_idx]:
                    if st.button(f"{sv}\n{sl}", key=f"sc_{key}_{sv}", use_container_width=True, type="primary" if cur_s == sv else "secondary"):
                        st.session_state.scores[key] = sv
                        st.rerun()
            
            st.session_state.remarks[key + "_remark"] = st.text_input("Remarks / Evidence", key=f"rem_{key}", value=st.session_state.remarks.get(key + "_remark", ""), placeholder="e.g. Completed MVP, 5 pilots...")
            st.divider()

    # ── Final Export ──
    st.markdown("### 📤 Finalize & Export")
    
    total_q = sum(len([f for d, dv in c["domains"].items() for f in dv["factors"]]) for c in COHORTS.values())
    total_f = len(st.session_state.scores)
    
    c1, c2 = st.columns([2, 1])
    with c1:
        st.write(f"**Total Progress:** {total_f} / {total_q} factors completed.")
        if total_f < total_q:
            st.warning("⚠️ You haven't completed all cohorts. Your Excel will contain partial data.")
        else:
            st.success("✅ All assessments completed!")
            
    with c2:
        if st.session_state.info["startup"] and st.session_state.info["founder"]:
            xlsx_data = build_xlsx(st.session_state.info, st.session_state.scores, st.session_state.remarks)
            st.download_button(
                label="📥 Download Assessment (.xlsx)",
                data=xlsx_data,
                file_name=f"{st.session_state.info['startup'].replace(' ','_')}_FRP_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        else:
            st.error("Enter Startup & Founder Name in sidebar to download.")

    st.markdown("""
    <div style="background:rgba(255,255,255,0.03); padding:1.5rem; border-radius:12px; margin-top:2rem; text-align:center;">
        <p style="color:#64748b; font-size:0.9rem;">Once downloaded, please send the <b>.xlsx</b> file to your assigned mentor or program coordinator.</p>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
